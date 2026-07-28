"""
VOC(사용자/운영자 질의응답 이력) 관리 API.
개별 등록/수정/삭제와, 엑셀(question/answer/department/resolved 컬럼) 일괄 업로드를 지원한다.
"""
import tempfile

import openpyxl
from fastapi import APIRouter, Depends, Form, UploadFile, File, HTTPException
from fastapi.concurrency import run_in_threadpool
from pydantic import BaseModel

from auth import require_admin
from cleaning import clean_text
from db import get_pool, embed_text, vector_literal
from server_files import read_upload_or_server_file
from spreadsheet import TABLE_EXTS, read_table_meta, load_table_rows
from uploads import (
    create_upload_session, get_upload_session, delete_upload_session, load_options,
)

router = APIRouter(prefix="/api/voc", tags=["voc"])

# 사내 VOC 엑셀 표준 포맷: 4행이 헤더, 이 4개 컬럼만 쓴다(의뢰번호/클러스터/의뢰자 등은 안 씀).
_VOC_HEADER_ROW = 4
_COL_REQUEST = "의뢰내용"
_COL_ACTION_DATE = "조치일"
_COL_RESOLUTION = "처리내용"
_COL_SATISFACTION = "만족도"
# 이 값이면 제외(불만족류). 매우만족/만족/보통/빈값은 그대로 사용.
_EXCLUDED_SATISFACTION = {"불만족", "매우불만족"}


class VocIn(BaseModel):
    question: str
    answer: str
    department: str | None = None
    resolved: bool = True


@router.get("")
async def list_voc(q: str | None = None, admin: str = Depends(require_admin)):
    pool = await get_pool("voc_db_dsn")
    if q:
        rows = await pool.fetch(
            "SELECT id, question, answer, department, resolved, created_at FROM voc_records "
            "WHERE question ILIKE '%' || $1 || '%' ORDER BY created_at DESC LIMIT 200",
            q,
        )
    else:
        rows = await pool.fetch(
            "SELECT id, question, answer, department, resolved, created_at FROM voc_records "
            "ORDER BY created_at DESC LIMIT 200"
        )
    return [dict(r) for r in rows]


@router.post("")
async def create_voc(body: VocIn, admin: str = Depends(require_admin)):
    vec = await embed_text(f"{body.question}\n{body.answer}")
    pool = await get_pool("voc_db_dsn")
    row_id = await pool.fetchval(
        """
        INSERT INTO voc_records (question, answer, department, resolved, embedding)
        VALUES ($1, $2, $3, $4, $5::vector) RETURNING id
        """,
        body.question,
        body.answer,
        body.department,
        body.resolved,
        vector_literal(vec),
    )
    return {"id": row_id}


@router.patch("/{voc_id}")
async def update_voc(voc_id: int, body: VocIn, admin: str = Depends(require_admin)):
    vec = await embed_text(f"{body.question}\n{body.answer}")
    pool = await get_pool("voc_db_dsn")
    row = await pool.fetchrow(
        """
        UPDATE voc_records SET question=$1, answer=$2, department=$3, resolved=$4, embedding=$5::vector
        WHERE id=$6 RETURNING id
        """,
        body.question,
        body.answer,
        body.department,
        body.resolved,
        vector_literal(vec),
        voc_id,
    )
    if not row:
        raise HTTPException(404, "VOC 기록을 찾을 수 없습니다.")
    return {"ok": True}


@router.delete("/{voc_id}")
async def delete_voc(voc_id: int, admin: str = Depends(require_admin)):
    pool = await get_pool("voc_db_dsn")
    await pool.execute("DELETE FROM voc_records WHERE id = $1", voc_id)
    return {"ok": True}


def _header_row(ws, row_num: int) -> list[str]:
    row = next(ws.iter_rows(min_row=row_num, max_row=row_num))
    return [str(c.value).strip() if c.value else "" for c in row]


async def _insert_voc(pool, question: str, answer: str, department: str | None, resolved: bool) -> bool:
    if not question or not answer:
        return False
    vec = await embed_text(f"{question}\n{answer}")
    await pool.execute(
        """
        INSERT INTO voc_records (question, answer, department, resolved, embedding)
        VALUES ($1, $2, $3, $4, $5::vector)
        """,
        question, answer, department, resolved, vector_literal(vec),
    )
    return True


async def _import_raw_format(ws, header: list[str], pool) -> tuple[int, int]:
    """사내 VOC 표준 엑셀: 4행 헤더, 의뢰내용/조치일/처리내용/만족도만 사용."""
    col_idx = {name: i for i, name in enumerate(header)}

    def cell(row, name):
        i = col_idx[name]
        return row[i] if i < len(row) else None

    inserted = skipped = 0
    for row in ws.iter_rows(min_row=_VOC_HEADER_ROW + 1, values_only=True):
        request_content = cell(row, _COL_REQUEST)
        action_date = cell(row, _COL_ACTION_DATE)
        resolution = cell(row, _COL_RESOLUTION)
        satisfaction = cell(row, _COL_SATISFACTION)

        if not request_content or not action_date or not resolution:
            skipped += 1
            continue
        if str(satisfaction).strip() in _EXCLUDED_SATISFACTION:
            skipped += 1
            continue

        question = clean_text(str(request_content))
        answer = clean_text(str(resolution))
        if await _insert_voc(pool, question, answer, None, True):
            inserted += 1
        else:
            skipped += 1
    return inserted, skipped


async def _import_simple_format(ws, header: list[str], pool) -> tuple[int, int]:
    """1행 헤더 Question/Answer(대소문자 무관) + department/resolved(선택), 2행부터 데이터.
    이미 정제된 텍스트로 취급하되 혹시 남은 HTML은 안전하게 걷어낸다."""
    lower_idx = {h.lower(): i for i, h in enumerate(header)}
    q_idx, a_idx = lower_idx["question"], lower_idx["answer"]
    dept_idx = lower_idx.get("department")
    resolved_idx = lower_idx.get("resolved")

    def cell(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    inserted = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        question_raw, answer_raw = cell(row, q_idx), cell(row, a_idx)
        if not question_raw or not answer_raw:
            skipped += 1
            continue
        department = str(cell(row, dept_idx)) if cell(row, dept_idx) else None
        resolved_val = cell(row, resolved_idx)
        resolved = (
            str(resolved_val).strip().upper() not in ("FALSE", "0", "N", "NO")
            if resolved_val is not None else True
        )

        question = clean_text(str(question_raw))
        answer = clean_text(str(answer_raw))
        if await _insert_voc(pool, question, answer, department, resolved):
            inserted += 1
        else:
            skipped += 1
    return inserted, skipped


@router.post("/import")
async def import_voc_excel(
    file: UploadFile | None = File(None),
    server_path: str | None = Form(None),
    admin: str = Depends(require_admin),
):
    """엑셀 형식을 자동으로 인식해서 등록한다. 지원하는 두 형식:
    (1) 1행 헤더 Question/Answer(대소문자 무관, department/resolved 선택) — 이미 정제된 데이터용.
    (2) 사내 VOC 표준 포맷 — 4행 헤더(의뢰내용/조치일/처리내용/만족도), 조치일·처리내용 있는 행만,
        만족도 불만족/매우불만족 제외, 본문은 HTML 태그만 벗기고 그대로 보존."""
    _, content, _ = await read_upload_or_server_file(file, server_path, {".xlsx"})
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    wb = openpyxl.load_workbook(tmp_path, read_only=True)
    ws = wb.active
    pool = await get_pool("voc_db_dsn")

    row1 = _header_row(ws, 1)
    if {"question", "answer"}.issubset({h.lower() for h in row1}):
        inserted, skipped = await _import_simple_format(ws, row1, pool)
    elif ws.max_row >= _VOC_HEADER_ROW and {
        _COL_REQUEST, _COL_ACTION_DATE, _COL_RESOLUTION, _COL_SATISFACTION
    }.issubset(set(_header_row(ws, _VOC_HEADER_ROW))):
        inserted, skipped = await _import_raw_format(ws, _header_row(ws, _VOC_HEADER_ROW), pool)
    else:
        raise HTTPException(
            422,
            "엑셀 형식을 인식하지 못했습니다. 지원 형식: "
            "(1) 1행 헤더 Question/Answer, 2행부터 데이터, 또는 "
            f"(2) {_VOC_HEADER_ROW}행 헤더에 {_COL_REQUEST}/{_COL_ACTION_DATE}/"
            f"{_COL_RESOLUTION}/{_COL_SATISFACTION} 컬럼이 모두 있는 사내 표준 포맷.",
        )

    return {"inserted": inserted, "skipped": skipped}


# ---------------------------------------------------------------- 열 매핑 업로드(형식 자유)
# 고정된 두 포맷(1행 Question/Answer · 4행 사내표준)만 받던 것을 대체한다.
# 헤더 행을 자동으로 찾고(제목 줄이 위에 몇 개 있어도 됨), 어떤 열을 무엇으로 쓸지 고르게 한다.
_DSN_VOC = "voc_db_dsn"


def _guess(columns: list[str], candidates: list[str]) -> str:
    for c in columns:
        low = c.lower()
        if any(k in low for k in candidates):
            return c
    return ""


@router.post("/excel/preview")
async def preview_voc_table(
    file: UploadFile | None = File(None),
    server_path: str | None = Form(None),
    header_row: int | None = Form(None),
    admin: str = Depends(require_admin),
):
    """엑셀/CSV의 헤더 행을 자동으로 찾아 열 목록과 샘플을 돌려준다.
    header_row를 주면 그 행(1-based, 엑셀에서 보이는 실제 행 번호)을 헤더로 강제한다."""
    ext, content, filename = await read_upload_or_server_file(file, server_path, TABLE_EXTS)
    upload_id = await create_upload_session(_DSN_VOC, admin, filename, ext, "voc_table", content, {})
    session = await get_upload_session(_DSN_VOC, upload_id, admin, "voc_table")
    try:
        sheet, header, sample, total, detected = await run_in_threadpool(
            read_table_meta, session["saved_path"], 5, header_row)
    except Exception as e:  # noqa: BLE001
        await delete_upload_session(_DSN_VOC, upload_id)
        raise HTTPException(422, f"파일을 읽을 수 없습니다: {e}")
    if not header:
        await delete_upload_session(_DSN_VOC, upload_id)
        raise HTTPException(422, "빈 파일입니다(표를 찾지 못했습니다).")

    return {
        "upload_id": upload_id, "filename": filename, "sheet": sheet,
        "columns": header, "sample_rows": sample, "total_rows": total,
        "header_row": detected,
        # 사내 표준 포맷이면 매핑을 미리 채워준다(그대로 등록만 누르면 되게).
        "suggest": {
            "question_column": _guess(header, ["의뢰내용", "question", "문의", "질문", "요청"]),
            "answer_column": _guess(header, ["처리내용", "answer", "답변", "조치", "회신"]),
            "department_column": _guess(header, ["department", "부서", "팀"]),
            "exclude_column": _guess(header, ["만족도", "satisfaction"]),
        },
    }


class VocTableCommitIn(BaseModel):
    """exclude_column/exclude_values: 특정 열의 값이 이 목록에 있으면 건너뛴다
    (사내 표준 포맷의 '만족도=불만족/매우불만족' 제외 규칙을 일반화한 것).
    require_columns: 이 열들이 비어 있는 행은 건너뛴다(예: 조치일이 없는 미처리 건)."""
    upload_id: str
    header_row: int | None = None
    question_column: str
    answer_column: str
    department_column: str | None = None
    exclude_column: str | None = None
    exclude_values: list[str] = []
    require_columns: list[str] = []


@router.post("/excel/commit")
async def commit_voc_table(body: VocTableCommitIn, admin: str = Depends(require_admin)):
    session = await get_upload_session(_DSN_VOC, body.upload_id, admin, "voc_table")

    def _build(path: str):
        header, col_idx, rows = load_table_rows(path, body.header_row)
        for label, col in (("질문", body.question_column), ("답변", body.answer_column)):
            if col not in col_idx:
                raise ValueError(f"{label} 열이 파일에 없습니다: {col}")
        for col in [body.department_column, body.exclude_column, *body.require_columns]:
            if col and col not in col_idx:
                raise ValueError(f"존재하지 않는 열입니다: {col}")

        def cell(row, col):
            if not col or col not in col_idx:
                return None
            v = row[col_idx[col]]
            return None if v is None else str(v).strip()

        excluded = {v.strip() for v in body.exclude_values if v.strip()}
        built, skipped = [], 0
        for row in rows:
            if any(not cell(row, c) for c in body.require_columns):
                skipped += 1
                continue
            if body.exclude_column and (cell(row, body.exclude_column) or "") in excluded:
                skipped += 1
                continue
            q, a = cell(row, body.question_column), cell(row, body.answer_column)
            if not q or not a:
                skipped += 1
                continue
            built.append((clean_text(q), clean_text(a), cell(row, body.department_column)))
        return built, skipped

    try:
        items, skipped = await run_in_threadpool(_build, session["saved_path"])
    except ValueError as e:
        raise HTTPException(422, str(e))
    finally:
        await delete_upload_session(_DSN_VOC, body.upload_id)

    if not items:
        raise HTTPException(422, "등록할 행이 없습니다. 열 선택과 제외 조건을 확인하세요.")

    pool = await get_pool(_DSN_VOC)
    inserted = 0
    for q, a, dept in items:
        if await _insert_voc(pool, q, a, dept, True):
            inserted += 1
        else:
            skipped += 1
    return {"inserted": inserted, "skipped": skipped, "total": len(items)}
