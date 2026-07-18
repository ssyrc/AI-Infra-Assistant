"""
매뉴얼(엑셀/워드/PPT) 업로드 -> 파싱 미리보기 -> 편집 -> 발행 -> 버전/롤백 API.

버전 모델: 같은 title로 재업로드하면 새 manual_files 행(version+1, status=draft)이 생긴다.
'발행(publish)'하면 해당 행의 미임베딩 청크만 임베딩하고 status=published로 바꾸며,
같은 title의 기존 published 행은 archived로 내려간다.
'롤백'은 과거 archived 버전을 다시 publish 호출하는 것과 동일하다(이미 임베딩되어 있어 즉시 반영).
"""
import os
import re
import uuid
import json
import tempfile
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from fastapi.concurrency import run_in_threadpool
from pydantic import BaseModel, Field

from auth import require_admin
from config_store import get_config
from db import get_pool, embed_text, vector_literal
from parser import parse_file, SUPPORTED_EXTS
from cleaning import clean_text, clean_options_from_dict, CleanOptions

router = APIRouter(prefix="/api/manuals", tags=["manuals"])

_TMP_DIR = tempfile.gettempdir()


@router.get("")
async def list_manuals(admin: str = Depends(require_admin)):
    pool = await get_pool("manual_db_dsn")
    rows = await pool.fetch(
        """
        SELECT id, title, filename, source_type, version, status, uploaded_by, uploaded_at, published_at
        FROM manual_files
        ORDER BY title, version DESC
        """
    )
    return [dict(r) for r in rows]


@router.get("/{manual_id}")
async def get_manual(manual_id: int, admin: str = Depends(require_admin)):
    pool = await get_pool("manual_db_dsn")
    file_row = await pool.fetchrow("SELECT * FROM manual_files WHERE id = $1", manual_id)
    if not file_row:
        raise HTTPException(404, "문서를 찾을 수 없습니다.")
    chunks = await pool.fetch(
        "SELECT id, seq, section_title, page_no, chunk_text, (embedding IS NOT NULL) AS embedded "
        "FROM manual_chunks WHERE manual_file_id = $1 ORDER BY seq",
        manual_id,
    )
    return {"file": dict(file_row), "chunks": [dict(c) for c in chunks]}


@router.get("/{manual_id}/versions")
async def list_versions(manual_id: int, admin: str = Depends(require_admin)):
    pool = await get_pool("manual_db_dsn")
    title_row = await pool.fetchrow("SELECT title FROM manual_files WHERE id = $1", manual_id)
    if not title_row:
        raise HTTPException(404, "문서를 찾을 수 없습니다.")
    rows = await pool.fetch(
        "SELECT id, version, status, uploaded_at, published_at FROM manual_files "
        "WHERE title = $1 ORDER BY version DESC",
        title_row["title"],
    )
    return [dict(r) for r in rows]


async def _insert_draft(title: str, filename: str, source_type: str, uploaded_by: str,
                        chunks: list) -> dict:
    """청크 리스트를 새 draft 버전으로 저장하는 공통 헬퍼. chunks는 (section_title, page_no, text) 튜플."""
    pool = await get_pool("manual_db_dsn")
    async with pool.acquire() as conn:
        async with conn.transaction():
            next_version = await conn.fetchval(
                "SELECT COALESCE(MAX(version), 0) + 1 FROM manual_files WHERE title = $1", title
            )
            file_id = await conn.fetchval(
                """
                INSERT INTO manual_files (title, filename, source_type, uploaded_by, version, status)
                VALUES ($1, $2, $3, $4, $5, 'draft')
                RETURNING id
                """,
                title, filename, source_type, uploaded_by, next_version,
            )
            for seq, (section_title, page_no, text) in enumerate(chunks):
                await conn.execute(
                    """
                    INSERT INTO manual_chunks (manual_file_id, seq, section_title, page_no, chunk_text)
                    VALUES ($1, $2, $3, $4, $5)
                    """,
                    file_id, seq, section_title, page_no, text,
                )
    return {"manual_file_id": file_id, "version": next_version, "chunk_count": len(chunks)}


async def _create_upload_session(owner: str, filename: str, ext: str, kind: str,
                                 content: bytes, options: dict) -> str:
    """업로드 파일을 서버가 정한 경로에 저장하고 세션을 DB에 기록한다.
    클라이언트는 upload_id만 알 뿐 경로·확장자·옵션을 결정하지 못한다."""
    pool = await get_pool("manual_db_dsn")
    try:
        ttl_min = int(await get_config("upload_session_ttl_minutes", "60"))
    except (TypeError, ValueError):
        ttl_min = 60

    upload_id = uuid.uuid4().hex
    saved_path = os.path.join(_TMP_DIR, f"upload-{upload_id}{ext}")
    with open(saved_path, "wb") as f:
        f.write(content)

    await pool.execute(
        """
        INSERT INTO upload_sessions (upload_id, owner, filename, ext, saved_path, kind, options, expires_at)
        VALUES ($1,$2,$3,$4,$5,$6,$7::jsonb, now() + ($8 || ' minutes')::interval)
        """,
        upload_id, owner, filename, ext, saved_path, kind,
        json.dumps(options, ensure_ascii=False), str(ttl_min),
    )
    await _cleanup_expired_sessions(pool)
    return upload_id


async def _get_upload_session(upload_id: str, owner: str, expected_kind: str) -> dict:
    """소유자·만료·종류를 검증하고 세션을 돌려준다."""
    if not re.fullmatch(r"[0-9a-f]{32}", upload_id or ""):
        raise HTTPException(400, "잘못된 upload_id 형식입니다.")
    pool = await get_pool("manual_db_dsn")
    row = await pool.fetchrow(
        "SELECT * FROM upload_sessions WHERE upload_id = $1", upload_id
    )
    if not row:
        raise HTTPException(404, "업로드 세션이 없거나 만료되었습니다. 다시 업로드하세요.")
    if row["owner"] != owner:
        raise HTTPException(403, "다른 사용자의 업로드 세션입니다.")
    if row["expires_at"] < datetime.now(timezone.utc):
        await _delete_upload_session(upload_id)
        raise HTTPException(404, "업로드 세션이 만료되었습니다. 다시 업로드하세요.")
    if row["kind"] != expected_kind:
        raise HTTPException(400, "업로드 종류가 일치하지 않습니다.")
    if not os.path.exists(row["saved_path"]):
        await _delete_upload_session(upload_id)
        raise HTTPException(404, "임시 파일이 정리되었습니다. 다시 업로드하세요.")
    return dict(row)


async def _delete_upload_session(upload_id: str):
    pool = await get_pool("manual_db_dsn")
    row = await pool.fetchrow("SELECT saved_path FROM upload_sessions WHERE upload_id = $1", upload_id)
    if row and os.path.exists(row["saved_path"]):
        try:
            os.unlink(row["saved_path"])
        except OSError:
            pass
    await pool.execute("DELETE FROM upload_sessions WHERE upload_id = $1", upload_id)


async def _cleanup_expired_sessions(pool):
    """만료된 미사용 preview 파일을 정리한다."""
    rows = await pool.fetch("SELECT upload_id, saved_path FROM upload_sessions WHERE expires_at < now()")
    for r in rows:
        if os.path.exists(r["saved_path"]):
            try:
                os.unlink(r["saved_path"])
            except OSError:
                pass
    if rows:
        await pool.execute("DELETE FROM upload_sessions WHERE expires_at < now()")


async def _read_upload(file: UploadFile, allowed_exts: set[str]) -> tuple[str, bytes]:
    """확장자·크기·빈 파일을 검증하고 내용을 읽는다."""
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in allowed_exts:
        raise HTTPException(422, f"지원하지 않는 형식입니다. 지원: {', '.join(sorted(allowed_exts))}")
    try:
        max_mb = int(await get_config("upload_max_mb", "50"))
    except (TypeError, ValueError):
        max_mb = 50
    limit = max_mb * 1024 * 1024

    content = await file.read(limit + 1)
    if len(content) > limit:
        raise HTTPException(413, f"파일이 너무 큽니다(최대 {max_mb}MB).")
    if not content:
        raise HTTPException(422, "빈 파일입니다.")

    # 컨테이너 계열(xlsx/docx/pptx)은 실제 zip인지 매직바이트로 확인
    if ext in (".xlsx", ".docx", ".pptx") and not content.startswith(b"PK"):
        raise HTTPException(422, "파일이 손상되었거나 형식이 올바르지 않습니다.")
    if ext == ".pdf" and not content.startswith(b"%PDF"):
        raise HTTPException(422, "PDF 파일이 손상되었거나 형식이 올바르지 않습니다.")
    return ext, content


@router.post("/preview")
async def preview_document(
    file: UploadFile = File(...),
    strip_html: bool = Form(True),
    collapse_space: bool = Form(True),
    drop_urls: bool = Form(False),
    include_speaker_notes: bool = Form(False),
    admin: str = Depends(require_admin),
):
    """docx/pptx/pdf/txt/md 전처리 미리보기. 아직 DB에 저장하지 않는다.
    정제 옵션은 서버가 세션에 저장하고, commit 때 그 옵션을 그대로 사용한다."""
    ext, content = await _read_upload(file, SUPPORTED_EXTS)
    options = {
        "strip_html": strip_html, "collapse_space": collapse_space,
        "drop_urls": drop_urls, "include_speaker_notes": include_speaker_notes,
    }
    upload_id = await _create_upload_session(admin, file.filename, ext, "document", content, options)
    session = await _get_upload_session(upload_id, admin, "document")

    opts = CleanOptions(strip_html=strip_html, collapse_space=collapse_space, drop_urls=drop_urls)
    try:
        chunks = await run_in_threadpool(
            parse_file, session["saved_path"], opts, include_speaker_notes)
    except Exception as e:  # noqa: BLE001
        await _delete_upload_session(upload_id)
        raise HTTPException(422, f"전처리 실패: {e}")

    if not chunks:
        await _delete_upload_session(upload_id)
        raise HTTPException(422, "문서에서 추출된 내용이 없습니다.")

    preview = [
        {"seq": i, "section_title": c.section_title, "page_no": c.page_no,
         "chunk_text": c.chunk_text, "char_count": len(c.chunk_text)}
        for i, c in enumerate(chunks[:50])
    ]
    return {"upload_id": upload_id, "filename": file.filename,
            "total_chunks": len(chunks), "preview_chunks": preview, "options": options}


class DocCommitIn(BaseModel):
    upload_id: str
    title: str = Field(min_length=1, max_length=200)


@router.post("/commit")
async def commit_document(body: DocCommitIn, uploaded_by: str = Depends(require_admin)):
    """preview에서 확인한 문서를 draft로 저장한다.
    확장자·경로·정제 옵션은 모두 서버 세션에서 가져오므로 미리보기 결과와 반드시 일치한다."""
    session = await _get_upload_session(body.upload_id, uploaded_by, "document")
    saved_options = session["options"] if isinstance(session["options"], dict) else json.loads(session["options"])
    opts = clean_options_from_dict(saved_options)

    try:
        parsed = await run_in_threadpool(
            parse_file, session["saved_path"], opts,
            bool(saved_options.get("include_speaker_notes", False)))
    finally:
        await _delete_upload_session(body.upload_id)

    if not parsed:
        raise HTTPException(422, "문서에서 추출된 내용이 없습니다.")
    chunks = [(c.section_title, c.page_no, c.chunk_text) for c in parsed]
    return await _insert_draft(body.title, session["filename"], "document", uploaded_by, chunks)


@router.post("/excel/preview")
async def preview_excel(
    file: UploadFile = File(...),
    strip_html: bool = Form(True),
    collapse_space: bool = Form(True),
    drop_urls: bool = Form(False),
    admin: str = Depends(require_admin),
):
    """엑셀 컬럼 목록과 샘플 행, 전체 행 수를 반환한다."""
    ext, content = await _read_upload(file, {".xlsx", ".xls"})
    options = {"strip_html": strip_html, "collapse_space": collapse_space, "drop_urls": drop_urls}
    upload_id = await _create_upload_session(admin, file.filename, ext, "spreadsheet", content, options)
    session = await _get_upload_session(upload_id, admin, "spreadsheet")

    def _read_header(path: str):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            wb.close()
            return None, None, [], 0
        header = [str(v).strip() if v is not None else f"column_{i}" for i, v in enumerate(header_row)]
        sample, total = [], 0
        for i, row in enumerate(it):
            total += 1
            if i < 5:
                sample.append([("" if v is None else str(v)) for v in row])
        sheet = ws.title
        wb.close()
        return sheet, header, sample, total

    try:
        sheet, header, sample, total = await run_in_threadpool(_read_header, session["saved_path"])
    except Exception as e:  # noqa: BLE001
        await _delete_upload_session(upload_id)
        raise HTTPException(422, f"엑셀을 읽을 수 없습니다: {e}")

    if not header:
        await _delete_upload_session(upload_id)
        raise HTTPException(422, "빈 엑셀 파일입니다.")

    return {"upload_id": upload_id, "filename": file.filename, "sheet": sheet,
            "columns": header, "sample_rows": sample, "total_rows": total, "options": options}


class ExcelCommitIn(BaseModel):
    upload_id: str
    title: str = Field(min_length=1, max_length=200)
    content_columns: list[str] = Field(min_length=1)
    title_column: str | None = None


@router.post("/excel/commit")
async def commit_excel(body: ExcelCommitIn, uploaded_by: str = Depends(require_admin)):
    """선택한 컬럼들로 행 단위 청크를 만들어 draft로 등록한다.
    정제 옵션은 preview 당시 서버에 저장된 값을 사용한다."""
    session = await _get_upload_session(body.upload_id, uploaded_by, "spreadsheet")
    saved_options = session["options"] if isinstance(session["options"], dict) else json.loads(session["options"])
    opts = clean_options_from_dict(saved_options)

    def _build(path: str):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(v).strip() if v is not None else f"column_{i}" for i, v in enumerate(next(it))]
        col_idx = {name: i for i, name in enumerate(header)}
        for c in body.content_columns:
            if c not in col_idx:
                wb.close()
                raise ValueError(f"존재하지 않는 컬럼입니다: {c}")

        built = []
        for row in it:
            if all(v is None for v in row):
                continue
            parts = []
            for c in body.content_columns:
                val = row[col_idx[c]]
                if val is None:
                    continue
                cleaned = clean_text(str(val), opts)
                if cleaned:
                    parts.append(f"{c}: {cleaned}")
            content = "\n".join(parts)
            if not content:
                continue
            section_title = None
            if body.title_column and body.title_column in col_idx:
                tv = row[col_idx[body.title_column]]
                section_title = clean_text(str(tv), opts) if tv is not None else None
            built.append((section_title or None, None, content))
        wb.close()
        return built

    try:
        chunks = await run_in_threadpool(_build, session["saved_path"])
    except ValueError as e:
        raise HTTPException(422, str(e))
    finally:
        await _delete_upload_session(body.upload_id)

    if not chunks:
        raise HTTPException(422, "선택한 컬럼으로 만들어진 내용이 없습니다.")
    return await _insert_draft(body.title, session["filename"], "spreadsheet", uploaded_by, chunks)


async def _assert_chunk_editable(pool, chunk_id: int) -> None:
    """청크의 부모 문서가 draft일 때만 편집을 허용한다.
    published/archived 문서를 수정하면 이미 서비스 중인 내용이 조용히 바뀌거나
    임베딩과 텍스트가 어긋나므로 금지한다."""
    row = await pool.fetchrow(
        """
        SELECT f.status FROM manual_chunks c
        JOIN manual_files f ON f.id = c.manual_file_id
        WHERE c.id = $1
        """,
        chunk_id,
    )
    if not row:
        raise HTTPException(404, "청크를 찾을 수 없습니다.")
    if row["status"] != "draft":
        raise HTTPException(
            409,
            f"'{row['status']}' 상태의 문서는 수정할 수 없습니다. "
            "같은 제목으로 새 버전을 업로드해 수정하세요.",
        )


@router.patch("/chunks/{chunk_id}")
async def update_chunk(chunk_id: int, chunk_text: str = Form(...), admin: str = Depends(require_admin)):
    """운영자가 자동 추출된 청크를 직접 교정한다. draft 상태에서만 허용된다."""
    pool = await get_pool("manual_db_dsn")
    await _assert_chunk_editable(pool, chunk_id)
    await pool.execute(
        "UPDATE manual_chunks SET chunk_text = $1, embedding = NULL WHERE id = $2",
        chunk_text, chunk_id,
    )
    return {"ok": True}


@router.delete("/chunks/{chunk_id}")
async def delete_chunk(chunk_id: int, admin: str = Depends(require_admin)):
    pool = await get_pool("manual_db_dsn")
    await _assert_chunk_editable(pool, chunk_id)
    await pool.execute("DELETE FROM manual_chunks WHERE id = $1", chunk_id)
    return {"ok": True}


@router.post("/{manual_id}/publish")
async def publish_manual(manual_id: int, admin: str = Depends(require_admin)):
    """draft/archived 버전을 발행(=검색 대상으로 전환)한다.
    archived 버전에 대해 호출하면 '롤백'과 동일하게 동작한다.

    동시성: 같은 title에 대해 두 사람이 동시에 발행하면 published가 2개가 될 수 있으므로
    title 기준 advisory lock으로 직렬화한다.
    임베딩이 하나라도 실패하면 상태를 바꾸지 않는다(부분 발행 방지)."""
    pool = await get_pool("manual_db_dsn")
    file_row = await pool.fetchrow("SELECT * FROM manual_files WHERE id = $1", manual_id)
    if not file_row:
        raise HTTPException(404, "문서를 찾을 수 없습니다.")
    if file_row["status"] == "published":
        return {"ok": True, "embedded_chunks": 0, "message": "이미 발행된 버전입니다."}

    embed_model = await get_config("vllm_embed_model", "bge-m3")
    try:
        embed_dim = int(await get_config("embed_dim", "1024"))
    except (TypeError, ValueError):
        embed_dim = 1024

    unembedded = await pool.fetch(
        "SELECT id, chunk_text FROM manual_chunks WHERE manual_file_id = $1 AND embedding IS NULL",
        manual_id,
    )
    embedded_count = 0
    for c in unembedded:
        try:
            vec = await embed_text(c["chunk_text"])
        except Exception as e:  # noqa: BLE001
            raise HTTPException(
                503,
                f"임베딩 서버 오류로 발행을 중단했습니다({embedded_count}/{len(unembedded)}개 완료). "
                f"서버 상태를 확인한 뒤 다시 발행하세요. 원인: {e}",
            )
        if len(vec) != embed_dim:
            raise HTTPException(
                500,
                f"임베딩 차원이 맞지 않습니다(모델 {len(vec)} vs 스키마 {embed_dim}). "
                "설정의 embed_dim과 임베딩 모델을 확인하세요.",
            )
        await pool.execute(
            "UPDATE manual_chunks SET embedding = $1::vector, embed_model = $2, embed_dim = $3 WHERE id = $4",
            vector_literal(vec), embed_model, embed_dim, c["id"],
        )
        embedded_count += 1

    async with pool.acquire() as conn:
        async with conn.transaction():
            # 같은 title 발행을 직렬화 (advisory lock, 트랜잭션 종료 시 자동 해제)
            await conn.execute("SELECT pg_advisory_xact_lock(hashtext($1))", file_row["title"])
            await conn.execute(
                "UPDATE manual_files SET status = 'archived' "
                "WHERE title = $1 AND status = 'published' AND id != $2",
                file_row["title"], manual_id,
            )
            await conn.execute(
                "UPDATE manual_files SET status = 'published', published_at = now() WHERE id = $1",
                manual_id,
            )
    return {"ok": True, "embedded_chunks": embedded_count}


@router.delete("/{manual_id}")
async def delete_manual(manual_id: int, admin: str = Depends(require_admin)):
    pool = await get_pool("manual_db_dsn")
    status_row = await pool.fetchrow("SELECT status FROM manual_files WHERE id = $1", manual_id)
    if not status_row:
        raise HTTPException(404, "문서를 찾을 수 없습니다.")
    if status_row["status"] == "published":
        raise HTTPException(400, "발행 중인 버전은 삭제할 수 없습니다. 먼저 다른 버전을 발행하세요.")
    await pool.execute("DELETE FROM manual_files WHERE id = $1", manual_id)
    return {"ok": True}
