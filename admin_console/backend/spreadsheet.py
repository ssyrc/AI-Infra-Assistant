"""
표 형식 파일(.xlsx/.xls/.csv) 읽기 공통 유틸.

여러 탭(매뉴얼·커맨드 등)이 같은 방식으로 헤더/샘플을 미리 보고, 선택한 열을 정제해
가져올 수 있도록 파싱을 한 곳에 모은다. 실제 정제(clean_text)는 호출부에서 수행한다.

CSV는 엑셀과 완전히 같은 "열 매핑" 흐름을 타도록 여기서 흡수한다(호출부는 확장자를 몰라도 된다).
사내에서 만든 CSV는 UTF-8(BOM 포함)이거나 CP949(엑셀 한글 기본)인 경우가 많아 둘 다 시도하고,
구분자도 쉼표/탭/세미콜론/파이프를 자동 판별한다.
"""
import csv
import os

import openpyxl

CSV_EXTS = {".csv"}
EXCEL_EXTS = {".xlsx", ".xls"}
TABLE_EXTS = EXCEL_EXTS | CSV_EXTS

_CSV_ENCODINGS = ("utf-8-sig", "cp949", "utf-8", "latin-1")


def _read_csv_all(path: str) -> list[list[str]]:
    """CSV 전체를 문자열 2차원 리스트로 읽는다. 인코딩/구분자를 자동으로 맞춘다."""
    last_err: Exception | None = None
    for enc in _CSV_ENCODINGS:
        try:
            with open(path, "r", newline="", encoding=enc) as f:
                sample = f.read(8192)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel        # 한 열짜리 파일 등 판별 실패 시 기본(쉼표)
                return [row for row in csv.reader(f, dialect)]
        except UnicodeDecodeError as e:
            last_err = e
            continue
    raise ValueError(
        f"CSV 인코딩을 인식할 수 없습니다(UTF-8 또는 CP949로 저장해 주세요): {last_err}")


def _is_csv(path: str) -> bool:
    return os.path.splitext(path)[1].lower() in CSV_EXTS


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def detect_header_row(rows: list[list], max_scan: int = 15) -> int:
    """헤더로 보이는 행의 인덱스(0-based)를 추정한다.

    엑셀마다 1행부터 표가 시작하기도 하고, 2행에 제목이 있고 4행부터 표가 나오기도 한다.
    "헤더 행은 (1) 채워진 칸이 여러 개고 (2) 각 칸이 짧은 라벨이며 (3) 값이 서로 다르고
    (4) 바로 아래에 비슷한 폭의 데이터 행이 이어진다"는 성질로 점수를 매겨 가장 높은 행을 고른다.
    """
    best_idx, best_score = 0, -1.0
    for i, row in enumerate(rows[:max_scan]):
        cells = [_norm(v) for v in row]
        filled = [c for c in cells if c]
        if len(filled) < 2:
            continue
        # 제목 줄(한 칸에 긴 문장)이 헤더로 뽑히지 않도록 긴 칸에 벌점
        long_cells = sum(1 for c in filled if len(c) > 40)
        distinct = len(set(filled)) / len(filled)
        # 아래로 3행이 이 폭을 유지하는지(데이터가 실제로 이어지는지)
        follow = 0
        for r in rows[i + 1:i + 4]:
            if sum(1 for v in r if _norm(v)) >= max(2, len(filled) * 0.6):
                follow += 1
        score = len(filled) + distinct * 2 + follow * 1.5 - long_cells * 3
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _header_of(row) -> list[str]:
    return [str(v).strip() if v is not None and str(v).strip() else f"column_{i}"
            for i, v in enumerate(row)]


def _all_rows(path: str) -> list[list]:
    """엑셀/CSV를 2차원 리스트로 읽는다. 빈 행도 그대로 둔다 —
    그래야 헤더 행 번호가 사용자가 엑셀에서 보는 실제 행 번호와 일치한다."""
    if _is_csv(path):
        return [list(r) for r in _read_csv_all(path)]
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return [list(r) for r in wb.active.iter_rows(values_only=True)]
    finally:
        wb.close()


def _sheet_name(path: str) -> str:
    if _is_csv(path):
        return "CSV"
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return wb.active.title
    finally:
        wb.close()


def read_table_meta(path: str, sample_size: int = 5, header_row: int | None = None):
    """(sheet, header, sample_rows, total_rows, header_row)를 반환한다.

    header_row(1-based)를 주지 않으면 자동으로 찾는다 — 1행부터 표가 시작하는 파일도,
    위에 제목/설명 줄이 몇 개 있고 중간부터 표가 시작하는 파일도 그대로 받기 위함이다.
    """
    rows = _all_rows(path)
    if not rows:
        return None, [], [], 0, 1
    idx = (header_row - 1) if header_row else detect_header_row(rows)
    idx = max(0, min(idx, len(rows) - 1))
    header = _header_of(rows[idx])
    body = [r for r in rows[idx + 1:] if any(_norm(v) for v in r)]
    sample = [[_norm(v) for v in r] for r in body[:sample_size]]
    return _sheet_name(path), header, sample, len(body), idx + 1


def load_table_rows(path: str, header_row: int | None = None):
    """(header, col_idx, rows)를 반환한다. col_idx는 열 이름 -> 인덱스 매핑이다.
    header_row(1-based)를 주지 않으면 read_table_meta와 같은 방식으로 자동 판별한다."""
    rows = _all_rows(path)
    if not rows:
        return [], {}, []
    idx = (header_row - 1) if header_row else detect_header_row(rows)
    idx = max(0, min(idx, len(rows) - 1))
    header = _header_of(rows[idx])
    col_idx = {name: i for i, name in enumerate(header)}
    # 헤더보다 열이 적은 행이 있어도 인덱스 접근이 터지지 않게 길이를 맞춘다.
    body = [r + [None] * (len(header) - len(r)) if len(r) < len(header) else r
            for r in rows[idx + 1:] if any(_norm(v) for v in r)]
    return header, col_idx, body


# 이전 이름 유지(호출부 점진 이행용).
read_excel_meta = read_table_meta
load_excel_rows = load_table_rows
